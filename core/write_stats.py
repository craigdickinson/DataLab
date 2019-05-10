"""
Created on 6 Sep 2016

@author: bowdenc
"""
import os.path
from datetime import datetime

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


class StatsOutput:
    """
    Methods to write statistics from LoggerStats to Excel.
    """

    def __init__(self, output_dir):
        """Create workbook object and delete initial worksheet (called Sheet)."""

        # Path to save output
        self.output_dir = output_dir

        # Dictionary to hold statistics data frames
        self.dict_stats = {}

        # Stats data frame for file export
        self.logger_id = ''
        self.stats = []
        self.start = []
        self.end = []
        self.df_stats = pd.DataFrame()

        self.wb = Workbook()
        sheet_name = self.wb.sheetnames[0]
        ws = self.wb[sheet_name]
        self.wb.remove(ws)

    def compile_stats_dataframe(self, logger, sample_start, sample_end, logger_stats):
        """
        Compile statistics into data frame for exporting and for use by gui.
        :param logger: object
        :param sample_start
        :param sample_end
        :param logger_stats: object
        :return: None
        """

        # Store logger id
        self.logger_id = logger.logger_id

        # Number of stats rows
        num_pts = len(logger_stats.min)

        # Create headers
        channels = logger.stats_channel_names
        # channel_header = [x for chan in channels for x in [chan] + ['', '', '']]
        channel_header = [x for chan in channels for x in [chan] * 4]

        stats_header = ['min', 'max', 'mean', 'std'] * len(channels)

        units = logger.stats_channel_units
        units_header = [x for unit in units for x in [unit] * 4]
        # CD: The above is equivalent to
        # units_header = []
        # for unit in units:
        #     for i in range(4):
        #         units_header.append(unit)

        mn = logger_stats.min
        mx = logger_stats.max
        ave = logger_stats.mean
        std = logger_stats.std

        # Order stats as:
        # [[channel1: min max ave std [0]... channelM: min max ave std [0],...
        #  [channel1: min max ave std [N]... channelM: min max ave std [N]]
        # zip creates tuple of: ((ch1_min, ch1_max, ch1_ave, ch1_std),...,(chM_min, chM_max, chM_ave, chM_std))
        # Logic: Loop each data point, create list, loop each channel in zip, loop each stat in channel, add to list
        self.stats = [[stat for channel in zip(mn[k], mx[k], ave[k], std[k]) for stat in channel]
                      for k in range(num_pts)]
        self.start = [datetime.strftime(dt, '%Y-%m-%d %H:%M:%S') for dt in sample_start]
        self.end = [datetime.strftime(dt, '%Y-%m-%d %H:%M:%S') for dt in sample_end]

        # CD: The above is equivalent to
        # ss = []
        # for k in range(num_pts):
        #     ss.append([])
        #     for stat in range(len(mn[0])):
        #         ss[k].append(mn[k][stat])
        #         ss[k].append(mx[k][stat])
        #         ss[k].append(ave[k][stat])
        #         ss[k].append(std[k][stat])
        #
        # print('break')
        # print(stats)
        # print(ss)

        # Create statistics data frame
        cols = pd.MultiIndex.from_arrays([channel_header, stats_header, units_header],
                                         names=['channels', 'stats', 'units'])
        df = pd.DataFrame(self.stats, index=sample_start, columns=cols)

        # Add stats data frame to logger dict for gui access
        self.dict_stats[logger.logger_id] = df

        # Now create a new data frame to write to file (csv/xlsx/hdf5)
        # Add End time column, reset index and rename first column to Start
        self.df_stats = pd.DataFrame(self.stats, index=self.start, columns=cols)
        self.df_stats.insert(loc=0, column='End', value=self.end)
        self.df_stats.reset_index(inplace=True)
        self.df_stats.rename({'index': 'Start'}, axis=1, inplace=True)

    def stats_to_hdf5(self):
        """Write stats to HDF5 file."""

        file_name = os.path.join(self.output_dir, 'Statistics.h5')
        self.df_stats.to_hdf(file_name, self.logger_id)

    def stats_to_csv(self):
        """Write stats to csv file."""

        file_name = os.path.join(self.output_dir, 'Statistics_' + self.logger_id + '.csv')
        self.df_stats.to_csv(file_name, index=False)

    def stats_to_excel(self, logger):
        """
        Write stats from a logger_stats object to Excel workbook.
        :param logger: object
        :return: None
        """

        # channel_header = list(self.stats_df.columns.unique(level='channels'))
        # stats_header = list(self.stats_df.columns.unique(level='stats'))
        # units_header = list(self.stats_df.columns.unique(level='units'))
        # stats = self.stats_df.values.tolist()

        # Create headers
        channels = logger.stats_channel_names
        channel_header = [x for chan in channels
                          for x in [chan] + ['', '', '']]

        stats_header = ['min', 'max', 'mean', 'std'] * len(channels)

        units = logger.stats_channel_units
        units_header = [x for unit in units
                        for x in [unit] * 4]

        # Data structure for Excel file - add in start and end time columns
        channel_header = ['Start', 'End'] + channel_header
        stats_header = ['', ''] + stats_header
        units_header = ['', ''] + units_header
        stats = [[self.start[k]] + [self.end[k]] + self.stats[k] for k in range(len(self.start))]

        # Create worksheet for logger
        ws = self.wb.create_sheet(title=self.logger_id)

        # Write headers
        ws.append(channel_header)
        ws.append(stats_header)
        ws.append(units_header)

        # Write data rows
        for row in stats:
            ws.append(row)

        # Formatting
        # Apply scientific format to data
        max_col = ws.max_column
        max_row = ws.max_row
        for row in range(4, max_row + 1):
            for col in range(3, max_col + 1):
                ws.cell(row=row, column=col).number_format = '0.00E+00'

        # Set width of date columns
        for col in range(1, 3):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def save_workbook(self):
        """Save workbook once all data has been written."""

        try:
            fname = 'Statistics.xlsx'
            self.wb.save(os.path.join(self.output_dir, fname))
        except:
            print('\n\nFailed to save ' + fname)
