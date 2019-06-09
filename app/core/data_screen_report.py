"""
Created on 22 Sep 2016

@author: bowdenc
"""

import os.path

from openpyxl import Workbook


class DataScreenReport(object):
    """Methods to write data screening results to Excel."""

    def __init__(self, project_name, campaign_name):
        """Constructor."""

        self.wb = Workbook()
        self.ws_summary = self.wb.active
        self.ws_summary.title = "Summary"
        self.project_name = project_name
        self.campaign_name = campaign_name
        self.bad_filenames = []
        self.bad_files = []

    def add_bad_filenames(self, logger_id, dict_bad_filenames):
        """Create list of logger filenames containing errors and add to summary list."""

        compiled_bad_files = self.compile_bad_files(logger_id, dict_bad_filenames)
        self.bad_filenames.extend(compiled_bad_files)

    def add_files_with_bad_data(self, logger_id, dict_bad_files):
        """Create list of logger files containing data errors and add to summary list."""

        compiled_bad_files = self.compile_bad_files(logger_id, dict_bad_files)
        self.bad_files.extend(compiled_bad_files)

    def compile_bad_files(self, logger_id, dict_bad_files):
        """
        Create list of errors encountered when processing logger filenames or screening files,
        in format to write to Excel.
        :param logger_id: Logger id
        :param dict_bad_files: Dictionary of filename-error pairs
        :return: Compiled list of [logger id, filename, error] for output to Excel
        """

        bad_files = [
            [logger_id] + [k, v]
            for k, v in zip(dict_bad_files.keys(), dict_bad_files.values())
        ]

        return bad_files

    def write_bad_filenames(self):
        """Write bad filenames to Data Screening Report workbook."""

        if len(self.bad_filenames) > 0:
            self.ws_bad_fname = self.wb.create_sheet(title="Bad Filenames")
            self.ws_bad_fname.append(["Logger ID", "File", "Error"])
            for row in self.bad_filenames:
                self.ws_bad_fname.append(row)

    def write_bad_files(self):
        """Write files with data errors to Data Screening Report workbook."""

        if len(self.bad_files) > 0:
            self.ws_bad_files = self.wb.create_sheet(title="Bad Files")
            self.ws_bad_files.append(["Logger ID", "File", "Error"])
            for row in self.bad_files:
                self.ws_bad_files.append(row)

    def save_workbook(self, path, fname):
        """Save workbook once all data has been written."""

        self.wb.save(os.path.join(path, fname))
