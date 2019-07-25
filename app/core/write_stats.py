"""
Class to compile and export logger stats.
"""
__author__ = "Craig Dickinson"

import os.path

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


class StatsOutput(object):
    """
    Methods to write statistics from LoggerStats to Excel.
    """

    def __init__(self, output_dir=""):
        """Create workbook object and delete initial worksheet (called Sheet)."""

        self.output_dir = output_dir
        self.logger_id = ""

        # List to hold stats data for all logger channels
        self.stats = []

        # Stats data frame for file export
        self.df_stats_export = pd.DataFrame()

        # Workbook object if writing stats to Excel
        self.wb = Workbook()
        sheet_name = self.wb.sheetnames[0]
        ws = self.wb[sheet_name]
        self.wb.remove(ws)

    def compile_stats(
        self, logger, sample_start, sample_end, logger_stats, logger_stats_filt
    ):
        """
        Compile statistics into data frame for exporting and for use by gui.
        :param logger: object
        :param sample_start
        :param sample_end
        :param logger_stats: object
        :param logger_stats_filt: object
        :return: df_stats
        """

        # Store logger id
        self.logger_id = logger.logger_id

        # Reorder the unfiltered logger stats
        stats_unfilt = self._reorder_stats(logger_stats)

        # Reorder the filtered logger stats (if processed)
        stats_filt = self._reorder_stats(logger_stats_filt)

        # Create headers
        channels = logger.channel_names
        channels_header_unfilt = [x for chan in channels for x in [chan] * 4]
        channels_header_filt = [
            x for chan in channels for x in [f"{chan} (filtered)"] * 4
        ]
        stats_header = ["min", "max", "mean", "std"] * len(channels)
        units_header = [x for unit in logger.channel_units for x in [unit] * 4]

        # If both unfiltered and filtered stats generated, join unfiltered and filtered stats columns
        if stats_unfilt and stats_filt:
            # Join unfiltered and filtered stats arrays
            self.stats = np.hstack((stats_unfilt, stats_filt))

            # Create headers containing unfiltered and filtered channels
            channels_header = channels_header_unfilt + channels_header_filt
            stats_header *= 2
            units_header *= 2
        # Filtered stats not generated
        elif stats_unfilt:
            self.stats = stats_unfilt
            channels_header = channels_header_unfilt
        # Unfiltered stats not generated
        elif stats_filt:
            self.stats = stats_filt
            channels_header = channels_header_filt
        # No stats exist - warn
        else:
            return pd.DataFrame()

        # Create pandas multi-index header
        header = self._create_header(channels_header, stats_header, units_header)

        # Create stats data frame for internal use
        df_stats = self._create_stats_dataframe(self.stats, sample_start, header)

        # Create an alternative stats data frame in a layout for writing to file (includes end timestamps column)
        self.df_stats_export = self._create_export_stats_dataframe(
            self.stats, sample_start, sample_end, header
        )

        # If unfiltered and filtered processed reorder stats data frame columns to
        # preferred order of (channel, channel (filtered)) pairs
        if stats_unfilt and stats_filt:
            cols = self._reorder_columns(df_stats)
            df_stats = df_stats[cols]
            self.df_stats_export = self.df_stats_export[["Start", "End"] + cols]

        return df_stats

    @staticmethod
    def _reorder_stats(logger_stats):
        """
        Order stats as:
            [[chan_1: min max ave std [0],..., chan_M: min max ave std [0]],
             [chan_1: min max ave std [N],..., chan_M: min max ave std [N]]].
        """

        if len(logger_stats.min) == 0:
            return []

        num_pts = len(logger_stats.min)

        # These arrays are lists of channels, e.g.
        # mn = [[chan_1_min[0],..., chan_M_min[0]],
        #       [chan_1_min[N],..., chan_M_min[N]]]
        mn = logger_stats.min
        mx = logger_stats.max
        ave = logger_stats.mean
        std = logger_stats.std

        # Zip creates tuple of: ((ch1_min, ch1_max, ch1_ave, ch1_std),...,(chM_min, chM_max, chM_ave, chM_std))
        # Logic: Loop each data point, create list, loop each channel in zip, loop each stat in channel, add to list
        stats = [
            [stat for channel in zip(mn[k], mx[k], ave[k], std[k]) for stat in channel]
            for k in range(num_pts)
        ]

        return stats

    @staticmethod
    def _create_header(channel_header, stats_header, units_header):
        """Create multi-index header of channel names, stats, and units to use in stats data frame."""

        if not len(channel_header) == len(stats_header) == len(units_header):
            raise ValueError(
                "Cannot create stats results header. Length of header rows is not equal."
            )

        header = pd.MultiIndex.from_arrays(
            [channel_header, stats_header, units_header],
            names=["channels", "stats", "units"],
        )
        return header

    @staticmethod
    def _create_stats_dataframe(stats, sample_start, header):
        return pd.DataFrame(data=stats, index=sample_start, columns=header)

    @staticmethod
    def _create_export_stats_dataframe(stats, sample_start, sample_end, header):
        """
        Create an alternative statistics data frame layout for exporting to file (csv/xlsx/hdf5).
        The only difference is that the sample end timestamps column is included and a standard integer index is used.
        """

        # Add End time column, reset index and rename first column to Start
        df = pd.DataFrame(data=stats, columns=header)
        df.insert(loc=0, column="Start", value=sample_start)
        df.insert(loc=1, column="End", value=sample_end)

        # Convert start and end columns to strings
        df["Start"] = df["Start"].dt.strftime("%Y-%m-%d %H:%M:%S")
        df["End"] = df["End"].dt.strftime("%Y-%m-%d %H:%M:%S")

        return df

    @staticmethod
    def _reorder_columns(df):
        channels = df.columns.unique(0)
        n = len(channels)
        channels_unfilt = channels[: n // 2]
        channels_filt = channels[n // 2 :]
        new_cols = [col for pair in zip(channels_unfilt, channels_filt) for col in pair]

        return new_cols

    def write_to_hdf5(self, mode="w"):
        """Write stats to HDF5 file."""

        # Create directory if does not exist
        self._ensure_dir_exists(self.output_dir)

        filename = "Statistics.h5"
        file_path = os.path.join(self.output_dir, filename)
        logger_id = replace_space_with_underscore(self.logger_id)
        self.df_stats_export.to_hdf(file_path, logger_id, mode=mode)

        return filename

    def write_to_csv(self):
        """Write stats to csv file."""

        # Create directory if does not exist
        self._ensure_dir_exists(self.output_dir)

        logger_id = replace_space_with_underscore(self.logger_id)
        filename = "Statistics_" + logger_id + ".csv"
        file_path = os.path.join(self.output_dir, filename)
        self.df_stats_export.to_csv(file_path, index=False)

        return filename

    def write_to_excel(self):
        """Write stats from a logger_stats object to Excel workbook."""

        # Convert headers and values as lists
        channels_header = self.df_stats_export.columns.unique(level=0).to_list()
        stats_header = self.df_stats_export.columns.get_level_values(level=1).to_list()
        units_header = self.df_stats_export.columns.get_level_values(level=2).to_list()
        data = self.df_stats_export.values.tolist()

        # Reformat channels header so as not to have repeating channels
        channels = channels_header[2:]
        channels = [x for chan in channels for x in [chan] + ["", "", ""]]
        channels_header = channels_header[:2] + channels

        # Create worksheet for logger
        logger_id = replace_space_with_underscore(self.logger_id)
        ws = self.wb.create_sheet(title=logger_id)

        # Write headers
        ws.append(channels_header)
        ws.append(stats_header)
        ws.append(units_header)

        # Write data rows
        for row in data:
            ws.append(row)

        # Formatting
        # Apply scientific format to data
        max_col = ws.max_column
        max_row = ws.max_row
        for row in range(4, max_row + 1):
            for col in range(3, max_col + 1):
                ws.cell(row=row, column=col).number_format = "0.00E+00"

        # Set width of date columns
        for col in range(1, 3):
            ws.column_dimensions[get_column_letter(col)].width = 20

    @staticmethod
    def _ensure_dir_exists(directory):
        """Create directory (and intermediate directories) if do not exist."""

        if directory != "" and os.path.exists(directory) is False:
            os.makedirs(directory)

    def save_workbook(self):
        """Save workbook once all worksheets have been created."""

        try:
            # Create directory if does not exist
            self._ensure_dir_exists(self.output_dir)

            filename = "Statistics.xlsx"
            file_path = os.path.join(self.output_dir, filename)
            self.wb.save(file_path)

            return filename
        except:
            print("\n\nFailed to save " + filename)


def replace_space_with_underscore(input_str):
    """Replace any spaces with underscores in string."""

    return "_".join(input_str.split(" "))
