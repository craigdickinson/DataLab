"""Stats screening module."""

__author__ = "Craig Dickinson"

import os.path

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from core.control import Control


class StatsScreening(object):
    """Class to perform statistical screening of loggers."""

    def __init__(self, control=Control()):
        self.control = control

        # Initialise logger stats objects
        self.stats_unfilt = LoggerStats()
        self.stats_filt = LoggerStats()

        # Stats writing object
        self.stats_out = StatsOutput(output_dir=self.control.stats_output_path)

        # To store stats for all datasets to load to gui
        self.dict_stats = {}

        # If writing stats HDF5 file, stats for all loggers are written to the same file
        # Set write mode to write new file for first logger then append for all others
        self.h5_write_mode = "w"
        self.h5_output_file_suffix = ""

    def init_logger_stats(self):
        """Set new stats objects for processing a new logger."""

        self.stats_unfilt = LoggerStats()
        self.stats_filt = LoggerStats()

    def file_stats_processing(self, df_file, data_screen, processed_file_num):
        """Stats processing module."""

        logger = data_screen.logger
        sample_length = data_screen.stats_sample_length
        df_stats = df_file.copy()
        df_stats_sample = pd.DataFrame()

        while len(df_stats) > 0:
            # Store the file number of processed sample (only of use for time step indexes)
            data_screen.stats_file_nums.append(processed_file_num)

            # Extract sample dataframe from main dataset
            df_stats_sample, df_stats = data_screen.sample_data(
                df_stats_sample, df_stats, sample_length, type="stats"
            )

            # Process sample if meets required length
            # TODO: Allowing short sample length (revisit)
            # if len(df_stats_sample) == sample_length:
            # if len(df_stats_sample) <= sample_length:
            # Unfiltered data
            if logger.process_type != "Filtered only":
                # Calculate sample stats
                self.stats_unfilt.calc_stats(df_stats_sample)
                data_screen.stats_processed = True

            # Filtered data
            if logger.process_type != "Unfiltered only":
                if data_screen.apply_filters is True:
                    # Apply low/high pass filtering
                    df_filt = data_screen.filter_data(df_stats_sample)

                    # Calculate sample stats
                    self.stats_filt.calc_stats(df_filt)
                    data_screen.stats_processed = True

            # Clear sample dataframe ready for next sample set
            df_stats_sample = pd.DataFrame()

        return data_screen.stats_processed

    def logger_stats_post(self, logger, data_screen):
        """
        Stats post-processing of all files for a given logger.
        Compile stats dataframe and export to file.
        """

        output_files = []

        # Create and store a dataframe of logger stats
        df_stats = self.stats_out.compile_stats(
            logger,
            data_screen.stats_file_nums,
            data_screen.stats_sample_start,
            data_screen.stats_sample_end,
            self.stats_unfilt,
            self.stats_filt,
        )

        if not df_stats.empty:
            # Store stats in dictionary for plotting in gui
            self.dict_stats[logger.logger_id] = df_stats

            # Export stats to requested file formats
            if self.control.stats_to_csv is True:
                stats_filename = self.stats_out.write_to_csv()

                # Add to output files list - to write to progress window
                rel_filepath = self.control.stats_output_folder + "/" + stats_filename
                output_files.append(rel_filepath)

            if self.control.stats_to_xlsx is True:
                self.stats_out.write_to_excel()

            if self.control.stats_to_h5 is True:
                stats_filename = self.stats_out.write_to_hdf5(self.h5_write_mode)

                # Add to output files list - to write to progress window
                rel_filepath = self.control.stats_output_folder + "/" + stats_filename
                output_files.append(rel_filepath + self.h5_output_file_suffix)

                # Set write mode to append to file for additional loggers
                self.h5_write_mode = "a"
                self.h5_output_file_suffix = " (appended)"

        return output_files

    def save_stats_excel(self):
        """Save stats workbook."""

        stats_filename = self.stats_out.save_workbook()
        output_file = self.control.stats_output_folder + "/" + stats_filename

        # Needs to be a list to update progress window
        return [output_file]


class LoggerStats(object):
    """Class to calculate and store logger statistics."""

    def __init__(self):
        """Lists for stats of each channel."""

        self.min = []
        self.max = []
        self.std = []
        self.mean = []

    def calc_stats(self, df_sample):
        """
        Calculate basic stats.
        Assumes at least two columns and first column is time.
        """

        data = df_sample[df_sample.columns[1:]]

        # Calculate min, max, mean and std for each channel
        mn = data.min()
        mx = data.max()
        ave = data.mean()
        std = data.std()

        # Append to internal list
        self.min.append(mn.values)
        self.max.append(mx.values)
        self.mean.append(ave.values)
        self.std.append(std.values)

        # TODO: McDermott project hack - don't keep this!
        # Hack for McDermott project to report slope between E and N time series instead of st. dev.
        # x = data.iloc[:, 1].values
        # y = data.iloc[:, 2].values
        # m = calc_slope(x, y)
        # self.std.append(np.array([m, m, m]))


def calc_slope(x, y):
    """Calculate the slope between two time series."""

    m, b = np.polyfit(x, y, 1)

    return m


class StatsOutput(object):
    """Class to compile and export logger stats."""

    def __init__(self, output_dir=""):
        """Create workbook object and delete initial worksheet (called Sheet)."""

        self.output_dir = output_dir
        self.logger_id = ""

        # List to hold stats data for all logger channels
        self.stats = []

        # Stats dataframe for file export
        self.df_stats_export = pd.DataFrame()

        # Workbook object if writing stats to Excel
        self.wb = Workbook()
        sheet_name = self.wb.sheetnames[0]
        ws = self.wb[sheet_name]
        self.wb.remove(ws)

    def compile_stats(
        self, logger, file_nums, sample_start, sample_end, logger_stats, logger_stats_filt
    ):
        """
        Compile statistics into dataframe for exporting and for use by gui.
        :param logger: object
        :param file_nums: Load case list assigned to each sample
        :param sample_start
        :param sample_end
        :param logger_stats: object
        :param logger_stats_filt: object
        :return: df_stats
        """

        # Store logger id
        self.logger_id = logger.logger_id

        # Reorder the unfiltered logger stats
        stats_unfilt = self._reorder_stats(logger_stats)

        # Reorder the filtered logger stats (if processed)
        stats_filt = self._reorder_stats(logger_stats_filt)

        # Create headers
        channels = logger.channel_names
        channels_header_unfilt = [x for chan in channels for x in [chan] * 4]
        channels_header_filt = [x for chan in channels for x in [f"{chan} (filtered)"] * 4]
        stats_header = ["min", "max", "mean", "std"] * len(channels)

        # TODO: McDermott project hack - don't keep this!
        # stats_header = ["min", "max", "mean", "E-N slope"] * len(channels)

        units_header = [x for unit in logger.channel_units for x in [unit] * 4]

        # If both unfiltered and filtered stats generated, join unfiltered and filtered stats columns
        if stats_unfilt and stats_filt:
            # Join unfiltered and filtered stats arrays
            self.stats = np.hstack((stats_unfilt, stats_filt))

            # Create headers containing unfiltered and filtered channels
            channels_header = channels_header_unfilt + channels_header_filt
            stats_header *= 2
            units_header *= 2
        # Filtered stats not generated
        elif stats_unfilt:
            self.stats = stats_unfilt
            channels_header = channels_header_unfilt
        # Unfiltered stats not generated
        elif stats_filt:
            self.stats = stats_filt
            channels_header = channels_header_filt
        # No stats exist - warn
        else:
            return pd.DataFrame()

        # Create pandas multi-index header
        header = self._create_header(channels_header, stats_header, units_header)

        # Create stats dataframe for internal use
        if logger.first_col_data == "Timestamp":
            df_stats = pd.DataFrame(data=self.stats, index=sample_start, columns=header)
            df_stats.index.name = "Date"
        else:
            df_stats = pd.DataFrame(data=self.stats, index=file_nums, columns=header)
            df_stats.index.name = "File Number"

        # Create an alternative stats dataframe in a layout for writing to file (includes end timestamps column)
        self.df_stats_export = self._create_export_stats_dataframe(
            self.stats, file_nums, sample_start, sample_end, header
        )

        # If unfiltered and filtered processed, reorder stats dataframe columns to
        # preferred order of (channel, channel (filtered)) pairs
        if stats_unfilt and stats_filt:
            cols = self._reorder_columns(df_stats)
            df_stats = df_stats[cols]
            self.df_stats_export = self.df_stats_export[["File Number", "Start", "End"] + cols]

        return df_stats

    @staticmethod
    def _reorder_stats(logger_stats):
        """
        Order stats as:
            [[chan_1: min max ave std [0],..., chan_M: min max ave std [0]],
             [chan_1: min max ave std [N],..., chan_M: min max ave std [N]]].
        """

        if len(logger_stats.min) == 0:
            return []

        num_pts = len(logger_stats.min)

        # These arrays are lists of channels, e.g.
        # mn = [[chan_1_min[0],..., chan_M_min[0]],
        #       [chan_1_min[N],..., chan_M_min[N]]]
        mn = logger_stats.min
        mx = logger_stats.max
        ave = logger_stats.mean
        std = logger_stats.std

        # Zip creates tuple of: ((ch1_min, ch1_max, ch1_ave, ch1_std),...,(chM_min, chM_max, chM_ave, chM_std))
        # Logic: Loop each data point, create list, loop each channel in zip, loop each stat in channel, add to list
        stats = [
            [stat for channel in zip(mn[k], mx[k], ave[k], std[k]) for stat in channel]
            for k in range(num_pts)
        ]

        return stats

    @staticmethod
    def _create_header(channel_header, stats_header, units_header):
        """Create multi-index header of channel names, stats, and units to use in stats dataframe."""

        if not len(channel_header) == len(stats_header) == len(units_header):
            raise ValueError(
                "Cannot create stats results header. Length of header rows is not equal."
            )

        header = pd.MultiIndex.from_arrays(
            [channel_header, stats_header, units_header], names=["channels", "stats", "units"]
        )
        return header

    @staticmethod
    def _create_export_stats_dataframe(stats, file_nums, sample_start, sample_end, header):
        """
        Create an alternative statistics dataframe layout for exporting to file (csv/xlsx/hdf5).
        The only difference is that the sample end timestamps column is included and a standard integer index is used.
        """

        # Add End time column, reset index and rename first column to Start
        df = pd.DataFrame(data=stats, columns=header)
        df.insert(loc=0, column="File Number", value=file_nums)
        df.insert(loc=1, column="Start", value=sample_start)
        df.insert(loc=2, column="End", value=sample_end)

        # Convert start and end column datetimes to strings for writing (if required)
        try:
            df["Start"] = df["Start"].dt.strftime("%Y-%m-%d %H:%M:%S")
            df["End"] = df["End"].dt.strftime("%Y-%m-%d %H:%M:%S")
        except:
            pass

        return df

    @staticmethod
    def _reorder_columns(df):
        channels = df.columns.unique(0)
        n = len(channels)
        channels_unfilt = channels[: n // 2]
        channels_filt = channels[n // 2 :]
        new_cols = [col for pair in zip(channels_unfilt, channels_filt) for col in pair]

        return new_cols

    def write_to_hdf5(self, mode="w"):
        """Write stats to HDF5 file."""

        filename = "Statistics.h5"
        file_path = os.path.join(self.output_dir, filename)
        logger_id = self.logger_id.replace(" ", "_")
        self.df_stats_export.to_hdf(file_path, logger_id, mode=mode)

        return filename

    def write_to_csv(self):
        """Write stats to csv file."""

        logger_id = self.logger_id.replace(" ", "_")
        filename = "Statistics_" + logger_id + ".csv"
        file_path = os.path.join(self.output_dir, filename)
        self.df_stats_export.to_csv(file_path, index=False)

        return filename

    def write_to_excel(self):
        """Write stats from a logger_stats object to Excel workbook."""

        # Convert headers and values as lists
        channels_header = self.df_stats_export.columns.unique(level=0).to_list()
        stats_header = self.df_stats_export.columns.get_level_values(level=1).to_list()
        units_header = self.df_stats_export.columns.get_level_values(level=2).to_list()
        data = self.df_stats_export.values.tolist()

        # Reformat channels header so as not to have repeating channels
        channels = channels_header[3:]
        channels = [x for chan in channels for x in [chan] + ["", "", ""]]
        channels_header = channels_header[:3] + channels

        # Create worksheet for logger
        logger_id = self.logger_id.replace(" ", "_")

        # Worksheet name length limit is 31
        if len(logger_id) > 31:
            logger_id = logger_id[:31]

        ws = self.wb.create_sheet(title=logger_id)

        # Write headers
        ws.append(channels_header)
        ws.append(stats_header)
        ws.append(units_header)

        # Write data rows
        for row in data:
            ws.append(row)

        # Formatting
        # Apply scientific format to data
        max_col = ws.max_column
        max_row = ws.max_row
        for row in range(4, max_row + 1):
            for col in range(4, max_col + 1):
                ws.cell(row, col).number_format = "0.00E+00"

        # Set width of date columns
        for col in range(2, 4):
            ws.column_dimensions[get_column_letter(col)].width = 19

    def save_workbook(self):
        """Save workbook once all worksheets have been created."""

        try:
            filename = "Statistics.xlsx"
            file_path = os.path.join(self.output_dir, filename)
            self.wb.save(file_path)

            return filename
        except Exception:
            print("\n\nFailed to save " + filename)
